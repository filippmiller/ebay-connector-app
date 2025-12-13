#!/usr/bin/env python3
"""Import bank XLSX exports as manual ledger transactions.

Goal
- Take transactions from one or more XLSX files (TD-style export) and insert them into the
  Accounting (ledger) tables.
- Each XLSX file becomes a "batch" (AccountingBankStatement) with source_type=MANUAL_XLSX.
- Each row becomes:
  - AccountingBankRow (for traceability / review)
  - AccountingTransaction (ledger) with source_type='manual' and source_id=<statement_id>

Idempotency
- We compute SHA256(file_bytes) and store it in accounting_bank_statement.file_hash.
- If a statement with the same file_hash already exists, we skip the file by default.
  Use --force to delete the previous import (statement + related rows + ledger txns) and re-import.

Usage (PowerShell)
  cd backend
  python scripts/import_manual_xlsx_ledger.py \
    --user-id <ADMIN_USER_UUID> \
    --files C:\\Users\\filip\\Downloads\\2019.xlsx C:\\Users\\filip\\Downloads\\2020.xlsx

Notes
- Requires DATABASE_URL in env (same as backend). Run from /backend so .env is picked up.
- Requires openpyxl (declared in backend/pyproject.toml).
"""

from __future__ import annotations

import argparse
import hashlib
import json
import os
import re
import sys
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any, Dict, Iterable, Iterator, Optional


# Add backend/ to sys.path for imports like other scripts do
SCRIPT_DIR = Path(__file__).resolve().parent
BACKEND_DIR = SCRIPT_DIR.parent
sys.path.insert(0, str(BACKEND_DIR))

# If run from repo root, app/config.py won't find backend/.env by default.
# Load it explicitly (no-op if DATABASE_URL already set).
try:
    from dotenv import load_dotenv  # type: ignore

    if not os.getenv("DATABASE_URL"):
        env_path = BACKEND_DIR / ".env"
        if env_path.exists():
            load_dotenv(env_path)
except Exception:
    # Missing python-dotenv is fine; env vars may already be present.
    pass

def _import_openpyxl():
    try:
        import openpyxl  # type: ignore

        return openpyxl
    except Exception as e:
        raise SystemExit(
            "openpyxl is required to parse XLSX. "
            "Install it in your backend environment (e.g. `cd backend; poetry install`). "
            f"Import error: {type(e).__name__}: {e}"
        )


def _sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        while True:
            chunk = f.read(1024 * 1024)
            if not chunk:
                break
            h.update(chunk)
    return h.hexdigest()


def _norm_header(s: Any) -> str:
    if s is None:
        return ""
    if not isinstance(s, str):
        s = str(s)
    s = s.strip().lower()
    s = re.sub(r"\s+", " ", s)
    return s


def _clean_amount(val: Any) -> Optional[Decimal]:
    if val is None:
        return None

    # openpyxl may give numeric types already
    if isinstance(val, (int, float, Decimal)):
        try:
            return Decimal(str(val)).quantize(Decimal("0.01"))
        except Exception:
            return None

    if isinstance(val, str):
        s = val.strip()
        if s == "":
            return None
        # Strip currency formatting like [$$]1,234.56
        s = re.sub(r"[^0-9.\-]", "", s)
        if s in ("", "-", "."):
            return None
        try:
            return Decimal(s).quantize(Decimal("0.01"))
        except Exception:
            return None

    return None


def _parse_date(val: Any) -> Optional[date]:
    if val is None:
        return None

    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val

    if isinstance(val, str):
        s = val.strip()
        if not s:
            return None
        for fmt in ("%Y-%m-%d", "%m/%d/%y", "%m/%d/%Y"):
            try:
                return datetime.strptime(s, fmt).date()
            except ValueError:
                pass

    # Last-resort: try pandas-like parsing without importing pandas
    try:
        return datetime.fromisoformat(str(val)).date()
    except Exception:
        return None


def _stringify_account_number(v: Any) -> Optional[str]:
    if v is None:
        return None
    if isinstance(v, (int, Decimal)):
        return str(v)
    if isinstance(v, float):
        # Excel may store as float; avoid scientific notation
        if v.is_integer():
            return str(int(v))
        return str(v)
    s = str(v).strip()
    if s == "":
        return None
    # remove trailing .0
    if re.fullmatch(r"\d+\.0", s):
        return s.split(".", 1)[0]
    return s


@dataclass
class ParsedXlsxRow:
    excel_row: int
    sheet_name: str
    posting_date: date
    bank_rtn: Optional[str]
    account_number: Optional[str]
    transaction_type_raw: Optional[str]
    description: str
    debit: Optional[Decimal]
    credit: Optional[Decimal]
    check_number: Optional[str]

    @property
    def signed_amount(self) -> Decimal:
        credit = self.credit or Decimal("0")
        debit = self.debit or Decimal("0")
        # Credits are positive, debits are negative
        return (credit - debit).quantize(Decimal("0.01"))

    @property
    def direction_bank(self) -> str:
        return "CREDIT" if self.signed_amount >= 0 else "DEBIT"

    @property
    def direction_ledger(self) -> str:
        return "in" if self.signed_amount >= 0 else "out"

    @property
    def amount_abs(self) -> Decimal:
        return abs(self.signed_amount).quantize(Decimal("0.01"))


def _find_header_row(sheet) -> Optional[int]:
    """Find the row number (1-indexed) where headers live."""
    required = {"date", "description"}
    candidates = []

    for r in range(1, min(30, sheet.max_row or 0) + 1):
        vals = [c.value for c in sheet[r]]
        norm = {_norm_header(v) for v in vals if v is not None}
        # accept either 'debit'/'credit' or similar
        if required.issubset(norm) and ("debit" in norm or "credit" in norm):
            candidates.append(r)

    if candidates:
        return candidates[0]

    # Fallback: assume row 2 is header if row 1 is some title
    if sheet.max_row and sheet.max_row >= 2:
        return 2
    return None


def _parse_xlsx(path: Path) -> list[ParsedXlsxRow]:
    openpyxl = _import_openpyxl()

    wb = openpyxl.load_workbook(filename=str(path), data_only=True)
    parsed: list[ParsedXlsxRow] = []

    for sheet in wb.worksheets:
        header_row = _find_header_row(sheet)
        if not header_row:
            continue

        header_cells = [c.value for c in sheet[header_row]]
        header_map: Dict[str, int] = {}
        for idx, h in enumerate(header_cells):
            key = _norm_header(h)
            if key:
                header_map[key] = idx

        # Helper to fetch a cell by canonical header name
        def get(row_vals: list[Any], *names: str) -> Any:
            for name in names:
                i = header_map.get(name)
                if i is not None and i < len(row_vals):
                    return row_vals[i]
            return None

        for r in range(header_row + 1, (sheet.max_row or header_row) + 1):
            row_vals = [c.value for c in sheet[r]]

            dt = _parse_date(get(row_vals, "date"))
            desc = get(row_vals, "description")

            # Skip empty rows
            if dt is None and (desc is None or str(desc).strip() == ""):
                continue

            if dt is None:
                # Can't store without date; skip but keep it obvious
                continue

            description = str(desc).strip() if desc is not None else ""
            if description == "":
                # Many exports always have description; but keep empty as allowed.
                description = ""

            bank_rtn = _stringify_account_number(get(row_vals, "bank rtn", "rtn", "bank routing", "routing"))
            account_number = _stringify_account_number(get(row_vals, "account number", "account"))
            txn_type = get(row_vals, "transaction type", "type")
            txn_type_raw = str(txn_type).strip() if txn_type is not None else None

            debit = _clean_amount(get(row_vals, "debit"))
            credit = _clean_amount(get(row_vals, "credit"))

            # Sometimes both are empty; skip
            if (debit is None or debit == 0) and (credit is None or credit == 0):
                continue

            check = get(row_vals, "check number", "check", "check #")
            check_number = str(check).strip() if check is not None and str(check).strip() != "" else None

            parsed.append(
                ParsedXlsxRow(
                    excel_row=r,
                    sheet_name=sheet.title,
                    posting_date=dt,
                    bank_rtn=bank_rtn,
                    account_number=account_number,
                    transaction_type_raw=txn_type_raw,
                    description=description,
                    debit=debit,
                    credit=credit,
                    check_number=check_number,
                )
            )

    return parsed


def _statement_hash(bank_code: str, account_number: str, period_start: date, period_end: date) -> str:
    key = f"{bank_code}|{account_number}|{period_start.isoformat()}|{period_end.isoformat()}"
    return hashlib.sha256(key.encode()).hexdigest()


def _storage_id_for_row(file_path: Path, sheet: str, excel_row: int) -> str:
    return f"manual_xlsx:{file_path.name}:{sheet}:{excel_row}"


def _chunked(it: list[Any], size: int) -> Iterator[list[Any]]:
    for i in range(0, len(it), size):
        yield it[i : i + size]


def main() -> None:
    parser = argparse.ArgumentParser(description="Import XLSX bank exports into accounting ledger as manual transactions")
    parser.add_argument("--user-id", required=True, help="Admin user UUID to set created_by_user_id")
    parser.add_argument("--files", nargs="+", required=True, help="One or more .xlsx files")
    parser.add_argument("--bank-name", default="TD Bank", help="Bank display name for batch statements")
    parser.add_argument("--bank-code", default="TD", help="Bank code for batch statements (e.g., TD)")
    parser.add_argument("--currency", default="USD", help="Currency code")
    parser.add_argument("--force", action="store_true", help="Delete existing statement(s) with same file_hash and re-import")
    parser.add_argument("--dry-run", action="store_true", help="Parse and print summary, but do not write to DB")
    args = parser.parse_args()

    user_id: str = args.user_id
    bank_name: str = args.bank_name
    bank_code: str = args.bank_code
    currency: str = args.currency

    files = [Path(f) for f in args.files]
    for f in files:
        if not f.exists():
            raise SystemExit(f"File not found: {f}")
        if f.suffix.lower() != ".xlsx":
            raise SystemExit(f"Not an .xlsx file: {f}")

    # Dry-run mode: parse XLSX and print summaries without touching the DB.
    if args.dry_run:
        for path in files:
            file_hash = _sha256_file(path)
            print(f"[PARSE] {path} (file_hash={file_hash[:12]}...)")
            rows = _parse_xlsx(path)
            if not rows:
                print(f"  [WARN] No rows parsed from {path.name} (skipping)")
                continue

            period_start = min(r.posting_date for r in rows)
            period_end = max(r.posting_date for r in rows)
            account_number = next((r.account_number for r in rows if r.account_number), None) or ""
            account_last4 = account_number[-4:] if len(account_number) >= 4 else (account_number or None)

            total_credit = sum((r.amount_abs for r in rows if r.direction_ledger == "in"), Decimal("0"))
            total_debit = sum((r.amount_abs for r in rows if r.direction_ledger == "out"), Decimal("0"))

            print(
                f"  rows={len(rows)} period={period_start.isoformat()}..{period_end.isoformat()} "
                f"credit={total_credit} debit={total_debit} account_last4={account_last4}"
            )

        print("Done (dry-run).")
        return

    # Import DB models lazily so env errors are clearer
    from app.models_sqlalchemy import SessionLocal
    from app.models_sqlalchemy.models import AccountingBankStatement, AccountingBankRow, AccountingTransaction

    db = SessionLocal()
    try:
        total_files_imported = 0
        total_rows_imported = 0

        for path in files:
            file_hash = _sha256_file(path)

            existing_stmt = (
                db.query(AccountingBankStatement)
                .filter(AccountingBankStatement.file_hash == file_hash)
                .order_by(AccountingBankStatement.id.desc())
                .first()
            )

            if existing_stmt and not args.force:
                print(f"[SKIP] {path.name}: already imported as statement_id={existing_stmt.id} (file_hash match)")
                continue

            if existing_stmt and args.force:
                stmt_id = existing_stmt.id
                print(f"[FORCE] Deleting previous import for {path.name}: statement_id={stmt_id}")
                # Delete ledger txns first (no cascade)
                deleted_txns = (
                    db.query(AccountingTransaction)
                    .filter(
                        AccountingTransaction.source_type == "manual",
                        AccountingTransaction.source_id == stmt_id,
                    )
                    .delete(synchronize_session=False)
                )
                # Delete statement (cascades rows)
                db.delete(existing_stmt)
                db.commit()
                print(f"        deleted ledger txns: {deleted_txns}")

            print(f"[PARSE] {path}")
            rows = _parse_xlsx(path)
            if not rows:
                print(f"  [WARN] No rows parsed from {path.name} (skipping)")
                continue

            period_start = min(r.posting_date for r in rows)
            period_end = max(r.posting_date for r in rows)

            # Determine account
            account_number = next((r.account_number for r in rows if r.account_number), None) or ""
            account_last4 = account_number[-4:] if len(account_number) >= 4 else (account_number or None)

            total_credit = sum((r.amount_abs for r in rows if r.direction_ledger == "in"), Decimal("0"))
            total_debit = sum((r.amount_abs for r in rows if r.direction_ledger == "out"), Decimal("0"))

            stmt_hash = _statement_hash(bank_code, account_number or "", period_start, period_end)

            print(
                f"  rows={len(rows)} period={period_start.isoformat()}..{period_end.isoformat()} "
                f"credit={total_credit} debit={total_debit} account_last4={account_last4}"
            )

            stmt = AccountingBankStatement(
                bank_name=bank_name,
                bank_code=bank_code,
                account_last4=account_last4,
                currency=currency,
                statement_period_start=period_start,
                statement_period_end=period_end,
                opening_balance=None,
                closing_balance=None,
                total_credit=total_credit,
                total_debit=total_debit,
                status="parsed",
                file_hash=file_hash,
                statement_hash=stmt_hash,
                source_type="MANUAL_XLSX",
                raw_json=None,
                created_by_user_id=user_id,
                updated_by_user_id=user_id,
            )
            db.add(stmt)
            db.flush()  # stmt.id

            statement_id = int(stmt.id)

            # Insert rows + ledger txns in chunks to keep memory/cost reasonable.
            CHUNK = 500
            inserted_rows_for_file = 0

            for chunk in _chunked(rows, CHUNK):
                bank_rows: list[AccountingBankRow] = []

                for r in chunk:
                    signed = r.signed_amount
                    raw_txn = {
                        "source": "manual_xlsx",
                        "source_file": path.name,
                        "sheet": r.sheet_name,
                        "excel_row": r.excel_row,
                        "date": r.posting_date.isoformat(),
                        "bank_rtn": r.bank_rtn,
                        "account_number": r.account_number,
                        "transaction_type_raw": r.transaction_type_raw,
                        "description": r.description,
                        "debit": str(r.debit) if r.debit is not None else None,
                        "credit": str(r.credit) if r.credit is not None else None,
                        "signed_amount": str(signed),
                        "check_number": r.check_number,
                    }

                    bank_row = AccountingBankRow(
                        bank_statement_id=statement_id,
                        row_index=r.excel_row,
                        operation_date=r.posting_date,
                        posting_date=r.posting_date,
                        description_raw=r.description or "",
                        description_clean=r.description or "",
                        amount=signed,
                        balance_after=None,
                        currency=currency,
                        parsed_status="manual_xlsx",
                        match_status="unmatched",
                        dedupe_key=None,
                        # v1.0-ish fields
                        bank_code=bank_code,
                        bank_section=None,
                        bank_subtype=r.transaction_type_raw,
                        direction=r.direction_bank,
                        accounting_group=None,
                        classification=None,
                        classification_status="UNKNOWN",
                        check_number=r.check_number,
                        raw_transaction_json=raw_txn,
                        created_by_user_id=user_id,
                        updated_by_user_id=user_id,
                    )
                    bank_rows.append(bank_row)

                db.add_all(bank_rows)
                db.flush()  # bank_row ids

                ledger_txns: list[AccountingTransaction] = []
                for bank_row, r in zip(bank_rows, chunk, strict=True):
                    storage_id = _storage_id_for_row(path, r.sheet_name, r.excel_row)

                    ledger_txns.append(
                        AccountingTransaction(
                            date=r.posting_date,
                            amount=r.amount_abs,
                            direction=r.direction_ledger,
                            source_type="manual",
                            source_id=statement_id,
                            bank_row_id=bank_row.id,
                            account_name=(f"{bank_name} ****{account_last4}" if account_last4 else bank_name),
                            account_id=r.account_number,
                            counterparty=None,
                            description=r.description,
                            expense_category_id=None,
                            subcategory=r.transaction_type_raw,
                            storage_id=storage_id,
                            linked_object_type=None,
                            linked_object_id=None,
                            is_personal=False,
                            is_internal_transfer=False,
                            created_by_user_id=user_id,
                            updated_by_user_id=user_id,
                        )
                    )

                db.add_all(ledger_txns)
                db.flush()

                inserted_rows_for_file += len(chunk)
                total_rows_imported += len(chunk)

            db.commit()
            total_files_imported += 1

            print(f"[OK] {path.name}: statement_id={statement_id} inserted_rows={inserted_rows_for_file}")

        print(f"Done. files_imported={total_files_imported} rows_imported={total_rows_imported}")

    finally:
        db.close()


if __name__ == "__main__":
    main()
