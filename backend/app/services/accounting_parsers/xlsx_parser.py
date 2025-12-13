import io
from typing import List, Dict, Any

def _normalize_header(name: str) -> str:
    """Normalize CSV/XLSX header names to improve auto-mapping.

    This makes the CSV/XLSX parser a bit more tolerant to different export
    formats (Google Sheets, various banks, etc.).
    """
    return " ".join(name.strip().split()).lower()

def parse_xlsx_bytes(file_bytes: bytes) -> List[Dict[str, Any]]:
    """Best-effort XLSX parser used for bank/ledger imports.

    We avoid overfitting to a single template and instead:
    - read the first sheet,
    - treat the first non-empty row as headers,
    - normalize header names similarly to CSV,
    - return a list of row dicts.
    """

    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as exc:  # pragma: no cover - import error is surfaced at runtime
        raise RuntimeError(
            "XLSX support requires the 'openpyxl' dependency; "
            "please ensure it is installed in the backend environment."
        ) from exc

    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active

    rows: List[Dict[str, Any]] = []
    header_row: List[str] = []

    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        vals = [cell for cell in row]
        # Skip completely empty rows at the top
        if not any(vals) and not header_row:
            continue
        if not header_row:
            header_row = [str(v) if v is not None else "" for v in vals]
            continue

        normalized: Dict[str, Any] = {}
        for col_idx, raw_val in enumerate(vals):
            if col_idx >= len(header_row):
                continue
            key = header_row[col_idx]
            if not isinstance(key, str) or not key.strip():
                continue
            key_norm = _normalize_header(key)
            normalized[key_norm] = raw_val

        if normalized:
            rows.append({"__index__": len(rows) + 1, **normalized})

    return rows
