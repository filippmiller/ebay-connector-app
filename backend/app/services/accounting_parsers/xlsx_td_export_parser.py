from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
import re
from typing import Any, Optional


@dataclass
class TdXlsxExportRow:
    sheet_name: str
    excel_row: int
    posting_date: date
    bank_rtn: Optional[str]
    account_number: Optional[str]
    transaction_type_raw: Optional[str]
    description: str
    debit: Optional[Decimal]
    credit: Optional[Decimal]
    check_number: Optional[str]

    @property
    def signed_amount(self) -> Decimal:
        credit = self.credit or Decimal("0")
        debit = self.debit or Decimal("0")
        return (credit - debit).quantize(Decimal("0.01"))

    @property
    def direction_bank(self) -> str:
        return "CREDIT" if self.signed_amount >= 0 else "DEBIT"

    @property
    def direction_ledger(self) -> str:
        return "in" if self.signed_amount >= 0 else "out"

    @property
    def amount_abs(self) -> Decimal:
        return abs(self.signed_amount).quantize(Decimal("0.01"))


def _norm_header(v: Any) -> str:
    if v is None:
        return ""
    s = str(v).strip().lower()
    s = re.sub(r"\s+", " ", s)
    return s


def _parse_date(v: Any) -> Optional[date]:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, str):
        s = v.strip()
        if not s:
            return None
        for fmt in ("%Y-%m-%d", "%m/%d/%y", "%m/%d/%Y"):
            try:
                return datetime.strptime(s, fmt).date()
            except ValueError:
                pass
    try:
        return datetime.fromisoformat(str(v)).date()
    except Exception:
        return None


def _clean_amount(v: Any) -> Optional[Decimal]:
    if v is None:
        return None
    if isinstance(v, Decimal):
        return v.quantize(Decimal("0.01"))
    if isinstance(v, (int, float)):
        try:
            return Decimal(str(v)).quantize(Decimal("0.01"))
        except Exception:
            return None
    if isinstance(v, str):
        s = v.strip()
        if s == "":
            return None
        s = re.sub(r"[^0-9.\-]", "", s)
        if s in ("", "-", "."):
            return None
        try:
            return Decimal(s).quantize(Decimal("0.01"))
        except Exception:
            return None
    return None


def _stringify_numeric(v: Any) -> Optional[str]:
    if v is None:
        return None
    if isinstance(v, int):
        return str(v)
    if isinstance(v, float):
        if v.is_integer():
            return str(int(v))
        return str(v)
    s = str(v).strip()
    if not s:
        return None
    if re.fullmatch(r"\d+\.0", s):
        return s.split(".", 1)[0]
    return s


def _find_header_row(sheet) -> Optional[int]:
    required = {"date", "description"}
    for r in range(1, min(30, sheet.max_row or 0) + 1):
        vals = [c.value for c in sheet[r]]
        norm = {_norm_header(x) for x in vals if x is not None}
        if required.issubset(norm) and ("debit" in norm or "credit" in norm):
            return r
    if sheet.max_row and sheet.max_row >= 2:
        return 2
    return None


def parse_td_xlsx_export(file_bytes: bytes) -> list[TdXlsxExportRow]:
    """Parse TD-style XLSX export into normalized rows.

    Expected columns (names vary slightly, case-insensitive):
    - Date
    - Bank RTN
    - Account Number
    - Transaction Type
    - Description
    - Debit
    - Credit
    - Check Number

    Returns rows across all worksheets.
    """
    try:
        import openpyxl  # type: ignore
    except Exception as e:
        raise RuntimeError(f"openpyxl is required to parse XLSX: {e}")

    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    out: list[TdXlsxExportRow] = []

    for sheet in wb.worksheets:
        header_row = _find_header_row(sheet)
        if not header_row:
            continue

        headers = [c.value for c in sheet[header_row]]
        idx: dict[str, int] = {}
        for i, h in enumerate(headers):
            k = _norm_header(h)
            if k:
                idx[k] = i

        def get(vals: list[Any], *names: str) -> Any:
            for name in names:
                j = idx.get(name)
                if j is not None and j < len(vals):
                    return vals[j]
            return None

        for r in range(header_row + 1, (sheet.max_row or header_row) + 1):
            vals = [c.value for c in sheet[r]]
            dt = _parse_date(get(vals, "date"))
            desc = get(vals, "description")

            if dt is None and (desc is None or str(desc).strip() == ""):
                continue
            if dt is None:
                continue

            debit = _clean_amount(get(vals, "debit"))
            credit = _clean_amount(get(vals, "credit"))
            if (debit is None or debit == 0) and (credit is None or credit == 0):
                continue

            out.append(
                TdXlsxExportRow(
                    sheet_name=sheet.title,
                    excel_row=r,
                    posting_date=dt,
                    bank_rtn=_stringify_numeric(get(vals, "bank rtn", "rtn", "routing", "bank routing")),
                    account_number=_stringify_numeric(get(vals, "account number", "account")),
                    transaction_type_raw=(
                        str(get(vals, "transaction type", "type")).strip()
                        if get(vals, "transaction type", "type") is not None
                        else None
                    ),
                    description=str(desc).strip() if desc is not None else "",
                    debit=debit,
                    credit=credit,
                    check_number=(
                        str(get(vals, "check number", "check", "check #")).strip()
                        if get(vals, "check number", "check", "check #") is not None
                        and str(get(vals, "check number", "check", "check #")).strip() != ""
                        else None
                    ),
                )
            )

    return out
